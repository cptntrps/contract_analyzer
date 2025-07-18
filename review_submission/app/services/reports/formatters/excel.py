"""
Excel Report Formatter

Generates professional Excel reports with styling and multiple sheets.
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from ....utils.logging.setup import get_logger

logger = get_logger(__name__)


class ExcelReportFormatter:
    """
    Excel report formatter for contract analysis results.
    
    Generates professional Excel workbooks with:
    - Changes table with color coding
    - Summary sheet with analysis metadata
    - Professional styling and formatting
    """
    
    def __init__(self):
        """Initialize Excel formatter"""
        pass
    
    def generate_changes_table(self, analysis_data: Dict[str, Any], output_path: str) -> str:
        """
        Generate Excel changes table with professional styling
        
        Args:
            analysis_data: Analysis results containing changes and metadata
            output_path: Path to save the Excel file
            
        Returns:
            Path to the generated Excel file
        """
        logger.info(f"Generating Excel changes table: {Path(output_path).name}")
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Changes Table"
        
        # Styling
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = [
            "Change #", "Section", "Change Type", "Classification", 
            "Template", "Document", "Relevance", 
            "Risk Level", "Recommendation"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data rows
        changes = analysis_data.get('analysis', [])
        for row, change in enumerate(changes, 2):
            row_data = [
                row - 1,  # Change number
                self._extract_section(change.get('deleted_text', '')),
                self._classify_change_type(change),
                change.get('classification', 'UNKNOWN'),
                self._truncate_text(change.get('deleted_text', ''), 100),
                self._truncate_text(change.get('inserted_text', ''), 100),
                change.get('explanation', 'No explanation provided'),
                self._get_change_risk_level(change),
                self._get_change_recommendation(change)
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                
                # Color coding based on classification
                if col == 4:  # Classification column
                    classification = change.get('classification', '').upper()
                    if classification == 'CRITICAL':
                        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                    elif classification == 'SIGNIFICANT':
                        cell.fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
        
        # Adjust column widths
        column_widths = [8, 15, 15, 15, 25, 25, 30, 12, 25]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Add summary sheet
        self._add_summary_sheet(wb, analysis_data)
        
        # Save workbook
        wb.save(output_path)
        logger.info(f"Excel file saved: {output_path}")
        
        return output_path
    
    def _add_summary_sheet(self, workbook: Workbook, analysis_data: Dict[str, Any]):
        """Add summary sheet to the workbook"""
        summary_ws = workbook.create_sheet("Summary")
        
        # Title
        title_cell = summary_ws.cell(row=1, column=1, value="Contract Analysis Summary")
        title_cell.font = Font(name='Arial', size=16, bold=True)
        
        # Summary data
        summary_data = [
            ["", ""],  # Empty row
            ["Contract", analysis_data.get('contract', 'Unknown')],
            ["Template", analysis_data.get('template', 'Unknown')],
            ["Analysis Date", analysis_data.get('date', datetime.now().strftime('%Y-%m-%d'))],
            ["Total Changes", str(analysis_data.get('changes', 0))],
            ["Similarity Score", f"{analysis_data.get('similarity', 0)}%"],
            ["Status", analysis_data.get('status', 'Completed')],
            ["", ""],  # Empty row
            ["Risk Analysis", ""],
            ["Critical Changes", str(len([c for c in analysis_data.get('analysis', []) if c.get('classification') == 'CRITICAL']))],
            ["Significant Changes", str(len([c for c in analysis_data.get('analysis', []) if c.get('classification') == 'SIGNIFICANT']))],
            ["Inconsequential Changes", str(len([c for c in analysis_data.get('analysis', []) if c.get('classification') == 'INCONSEQUENTIAL']))],
            ["Overall Risk Level", self._determine_risk_level(analysis_data)]
        ]
        
        for row, (label, value) in enumerate(summary_data, 2):
            if label:  # Skip empty rows
                label_cell = summary_ws.cell(row=row, column=1, value=label)
                label_cell.font = Font(bold=True)
                value_cell = summary_ws.cell(row=row, column=2, value=value)
                
                # Color code risk level
                if label == "Overall Risk Level":
                    if value == "HIGH":
                        value_cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                    elif value == "MEDIUM":
                        value_cell.fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
                    else:
                        value_cell.fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
        
        # Adjust column widths
        summary_ws.column_dimensions['A'].width = 20
        summary_ws.column_dimensions['B'].width = 30
    
    def _extract_section(self, text: str) -> str:
        """Extract section identifier from text"""
        if not text:
            return "Unknown"
        
        # Look for common section patterns
        words = text.split()[:5]  # First 5 words
        return ' '.join(words) if words else "Unknown"
    
    def _classify_change_type(self, change: Dict[str, Any]) -> str:
        """Classify the type of change"""
        deleted = change.get('deleted_text', '').strip()
        inserted = change.get('inserted_text', '').strip()
        
        if not deleted and inserted:
            return "Addition"
        elif deleted and not inserted:
            return "Deletion"
        elif deleted and inserted:
            return "Modification"
        else:
            return "Unknown"
    
    def _get_change_risk_level(self, change: Dict[str, Any]) -> str:
        """Get risk level for individual change"""
        classification = change.get('classification', '').upper()
        
        if classification == 'CRITICAL':
            return "High"
        elif classification == 'SIGNIFICANT':
            return "Medium"
        else:
            return "Low"
    
    def _get_change_recommendation(self, change: Dict[str, Any]) -> str:
        """Get recommendation for individual change"""
        classification = change.get('classification', '').upper()
        
        if classification == 'CRITICAL':
            return "Immediate legal review required"
        elif classification == 'SIGNIFICANT':
            return "Legal review recommended"
        else:
            return "Standard review process"
    
    def _truncate_text(self, text: str, max_length: int) -> str:
        """Truncate text to specified length with ellipsis"""
        if not text:
            return ""
        
        if len(text) <= max_length:
            return text
        
        return text[:max_length] + "..."
    
    def _determine_risk_level(self, analysis_data: Dict[str, Any]) -> str:
        """
        Determine overall risk level using minimum criticality rule
        
        Logic:
        - Any critical change = HIGH risk
        - Multiple significant changes = HIGH risk
        - Some significant changes = MEDIUM risk
        - Only inconsequential = LOW risk
        """
        changes = analysis_data.get('analysis', [])
        critical_count = len([c for c in changes if c.get('classification') == 'CRITICAL'])
        significant_count = len([c for c in changes if c.get('classification') == 'SIGNIFICANT'])
        
        # Minimum criticality rule: One critical change makes entire contract critical
        if critical_count > 0:
            return "HIGH"
        elif significant_count > 5:
            return "HIGH"
        elif significant_count > 0:
            return "MEDIUM"
        else:
            return "LOW"


__all__ = ['ExcelReportFormatter']